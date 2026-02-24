"""
Excel Reporter Module
Generates a styled Excel report of keyword classification results.

Output columns:
  document_name | classification | keyword | pages | x1 | y1 | x2 | y2 | width | height | confidence | text | source_image
"""

import os
from typing import List, Dict, Any


def save_classification_excel(
    matches: List[Dict[str, Any]],
    output_path: str,
    summary_sheet: bool = True,
) -> str:
    """
    Save keyword match results to a styled Excel workbook.

    Sheets created:
        1. "classification_results"  — detail rows, one per match
        2. "summary"                 — grouped view: doc → keyword → pages (optional)

    Args:
        matches:       Flat list of match dicts from keyword_classifier
        output_path:   Path to save the .xlsx file
        summary_sheet: Whether to also create the grouped summary sheet

    Returns:
        Absolute path to the saved Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    # ── Colour palette ──────────────────────────────────────────────────────
    HEADER_BG   = "1F4E79"   # dark blue
    HEADER_FG   = "FFFFFF"   # white
    ALT_ROW_BG  = "D6E4F0"   # light blue
    SUMMARY_BG  = "2E7D32"   # dark green
    SUMMARY_FG  = "FFFFFF"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = Workbook()

    # ── Sheet 1: Detail results ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "classification_results"

    detail_headers = [
        "document_name", "classification", "keyword",
        "page", "x1", "y1", "x2", "y2", "width", "height",
        "confidence", "text", "source_image",
    ]
    ws.append(detail_headers)

    # Style header row
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FG, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")

    for row_idx, m in enumerate(matches, start=2):
        x1 = int(m.get("x1", 0))
        y1 = int(m.get("y1", 0))
        x2 = int(m.get("x2", 0))
        y2 = int(m.get("y2", 0))
        width  = x2 - x1
        height = y2 - y1

        row_data = [
            m.get("document", ""),
            m.get("classification", ""),
            m.get("keyword", ""),
            m.get("page", ""),
            x1, y1, x2, y2,
            width, height,
            round(float(m.get("confidence", 0)), 4),
            m.get("text", ""),
            m.get("source_image", ""),
        ]
        ws.append(row_data)

        # Alternate row shading
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

        for cell in ws[row_idx]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-fit column widths
    col_min_widths = {
        1: 22, 2: 18, 3: 22, 4: 6,
        5: 7, 6: 7, 7: 7, 8: 7, 9: 8, 10: 8,
        11: 11, 12: 35, 13: 28,
    }
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        min_w = col_min_widths.get(col_idx, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, min_w)

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    if summary_sheet and matches:
        from keyword_classifier import group_matches_by_document

        ws2 = wb.create_sheet("summary")
        summary_headers = ["document_name", "classification", "keyword", "pages", "match_count"]
        ws2.append(summary_headers)

        sum_fill = PatternFill(start_color=SUMMARY_BG, end_color=SUMMARY_BG, fill_type="solid")
        sum_font = Font(bold=True, color=SUMMARY_FG, size=10)
        for cell in ws2[1]:
            cell.fill = sum_fill
            cell.font = sum_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws2.row_dimensions[1].height = 24

        grouped = group_matches_by_document(matches)
        sum_row = 2
        for doc, kw_dict in grouped.items():
            for kw, info in kw_dict.items():
                pages_str = ", ".join(str(p) for p in sorted(set(info["pages"])))
                classification = info["matches"][0].get("classification", kw.title())
                ws2.append([doc, classification, kw, pages_str, len(info["matches"])])
                for cell in ws2[sum_row]:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                if sum_row % 2 == 0:
                    for cell in ws2[sum_row]:
                        cell.fill = alt_fill
                sum_row += 1

        for col_idx, col_cells in enumerate(ws2.columns, start=1):
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=0,
            )
            ws2.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 14)

        ws2.freeze_panes = "A2"

    # ── Save ────────────────────────────────────────────────────────────────
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"[OK] Excel report saved: {output_path}  ({len(matches)} match rows)")
    return os.path.abspath(output_path)
